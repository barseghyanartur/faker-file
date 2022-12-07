import json
import os
from typing import Dict, List, Optional, Sequence, Union

from faker import Faker
from faker.providers import BaseProvider
from openpyxl import Workbook

from ..base import DEFAULT_REL_PATH, FileMixin, StringValue

__author__ = "Artur Barseghyan <artur.barseghyan@gmail.com>"
__copyright__ = "2022 Artur Barseghyan"
__license__ = "MIT"
__all__ = ("XlsxFileProvider",)


FAKER = Faker()


class XlsxFileProvider(BaseProvider, FileMixin):
    """CSV file provider.

    Usage example:

        from faker_file.providers.xlsx_file import XlsxFileProvider

        file = XlsxFileProvider(None).xlsx_file()

    Usage example with options:

        from faker_file.providers.xlsx_file import XlsxFileProvider

        file = XlsxFileProvider(None).xlsx_file(
            prefix="zzz",
            num_rows=100,
            data_columns={
                "name": "{{name}}",
                "residency": "{{address}}",
            },
            include_row_ids=True,
        )
    """

    extension: str = "xlsx"

    def xlsx_file(
        self,
        root_path: str = None,
        rel_path: str = DEFAULT_REL_PATH,
        prefix: Optional[str] = None,
        data_columns: Dict[str, str] = None,
        num_rows: int = 10,
        content: Optional[str] = None,
        **kwargs,
    ) -> StringValue:
        """Generate a XLSX file with random text.

        :param root_path: Path of your files root directory (in case of Django
            it would be `settings.MEDIA_ROOT`).
        :param rel_path: Relative path (from root directory).
        :param data_columns: The ``data_columns`` argument expects a list or a
            tuple of string tokens, and these string tokens will be passed to
            :meth:`pystr_format()
            <faker.providers.python.Provider.pystr_format>`
            for data generation. Argument Groups are used to pass arguments
            to the provider methods. Both ``header`` and ``data_columns`` must
            be of the same length.
        :param num_rows: The ``num_rows`` argument controls how many rows of
            data to generate, and the ``include_row_ids`` argument may be set
            to ``True`` to include a sequential row ID column.
        :param prefix: File name prefix.
        :param content: List of dicts with content (JSON-like format).
            If given, used as is.
        :return: Relative path (from root directory) of the generated file.
        """
        # Generic
        file_name = self._generate_filename(
            root_path=root_path,
            rel_path=rel_path,
            prefix=prefix,
        )

        # Specific
        if content is None:
            default_data_columns = {
                "name": "{{name}}",
                "residency": "{{address}}",
            }
            data_columns: Union[List, Dict] = (
                data_columns if data_columns else default_data_columns
            )
            content = FAKER.json(
                data_columns=data_columns,
                num_rows=num_rows,
            )
            content = json.loads(content)

        workbook = Workbook()
        worksheet = workbook.active
        keys = []

        for i in range(len(content)):
            sub_obj = content[i]
            if i == 0:
                keys = list(sub_obj.keys())
                for k in range(len(keys)):
                    worksheet.cell(row=(i + 1), column=(k + 1), value=keys[k])
            for j in range(len(keys)):
                worksheet.cell(
                    row=(i + 2), column=(j + 1), value=sub_obj[keys[j]]
                )

        workbook.save(file_name)

        # Generic
        file_name = StringValue(os.path.relpath(file_name, root_path))
        file_name.data = {"content": content}
        return file_name
